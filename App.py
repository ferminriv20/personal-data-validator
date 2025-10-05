import streamlit as st
import pandas as pd
import re
import json
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from validador import *


st.set_page_config(
    page_title='Personal Data Validator',
    page_icon= '✅', #Este icono se muestra  en la pestaña del navegador
)

st.markdown("""
<style>
    .main {
        background: #020617;
    }
    
    h1, .stTitle {
        color: #06b6d4 !important;
    }
    
    body, .stApp, .main, .css-1v0mbdj, .css-1c7y2kd, .css-1d391kg, .css-1lcbmhc, .css-1y4p8pa, .css-1v3fvcr, .css-1q8dd3e, .css-1c7y2kd, .stTextInput > label, .stSelectbox > label, .stButton > button, .stDownloadButton > button, .stDataFrame, .stTable, .stMarkdown, .stHeader, .stSubheader, .stCaption, .stText, .stMetric, .stAlert {
         color: #e2e8f0 !important;
    }
    
    .stButton > button, .stDownloadButton > button {
        background: linear-gradient(135deg, #06b6d4 0%, #0891b2 100%);
        color: white !important;
        border: none;
        border-radius: 12px;
        box-shadow: 0 15px 35px rgba(6, 182, 212, 0.4);
    }
     
    .stSelectbox > div > div {
        background: rgba(2, 6, 23, 0.8);
        border: 2px solid #06b6d4;
        color: #67e8f9;
    }
    .stSelectbox > div > div:hover {
        border: 2px solid #0891b2;
    }
    div[data-testid="stAlert"] {
        background-color: #0c1426 !important;
        color: #67e8f9 !important;
        border-radius: 12px !important;
        border: none !important;
        box-shadow: 0 15px 35px rgba(6, 182, 212, 0.4) !important;
    }
</style>
""", unsafe_allow_html=True)

st.title('🔍 Validador de Datos Personales')
st.markdown(""" ### Revisa la calidad de tus datos Fácilmente. """)


uploaded_file = st.file_uploader('📂 Sube tu archivo Excel', type=['xlsx'])
use_example = st.checkbox("📊 Usa un archivo de ejemplo", help="Selecciona esta opción para probar la aplicación con datos de muestra.")

if use_example:
    # Datos de ejemplo
    df = pd.DataFrame({
    'Nombre': [
        'Adriana', 'Carlota', 'Sergio', 'Débora', 'Cirino', 
        'ferm7n', 'Leocadia', 'Patricia', None, 'Arsenio',
        'Carmen', None, 'Rosalva', 'Evaristo', 'Marisela',
        'Teresa', 'tter3sa', 'Graciana', None, 'Soledad'
    ],
    'Apellido': [
        'Baños', 'tor33s', 'Cuadrado', 'Huertas', 'Casals',
        'Sosa', 'Verdú', 'Quero', 'Escolano', 'Andrés',
        'Ríos huertas', 'Pardo', 'Barranco', None, 'Alarcón',
        'Infante', 'lassarte:', 'Escalona', 'Martínez', 'Morán'
    ],
    'Teléfono': [
        '3955824818', '3485972605', '3125826914', '3405065262', '3440088174',
        '56789', '3312347313', '3255055540', '56789', '3017416589',
        '3537105736', '3740690501', '3884636047', '30030030025', '56789',
        '3150400887', '3574439046', '34352759395', '3207223962', '3833689323'
    ],
    'Correo': [
        'lquiroga@cespedes.es', 'bertanunez@yahoo.com', 'cloenicolas@gmail.com', 
        'larranagaeusebia@acuna.com', 'correo_invalido', 'celestinaruiz@gomez-recio.es',
        'tsarabia@hervia-roma.com', None, 'pepitapriego@hotmail.com', 'uaguirre@gmail.com',
        'saturnina63@hotmail.com', 'emilia54@gmail.com', 'rdaza@gmail.com', 
        'barcomaria-luisa@olivares.es', 'domingorius@hotmail.com', 'ferxxoxix316@@gmail.com',
        'doloressala@alberto.es', 'franciscojerez@yahoo.com', 'remedios57@hotmail.com', 
        'carriondimas@hotmail.com'
    ],
    'Tipo de documento': [
        'TI', 'CC', 'TI', 'TI', 'CC',
        'CC', 'XX', 'TI', 'TI', 'CC',
        'CC', None, 'CC', 'XX', 'CC',
        'CC', 'CC', 'CC', 'TI', 'TI'
    ],
    'ID': [
        '1411660738', 'abc123', '1375801277', '1196843017', '1105523456',
        '1094975081', '1750222033', '1224554818', '1228853027', '1294765429',
        None, '1974009467', '1117312531', '1315277521', '1980118600',
        '1908755906', '1761389500', '1857273988', '1.68853E+12', '1437720783'
    ],
    'Fecha de cumpleaños': [
        '1943-15-13', '2020-07-12', '1944-11-24', '1966-04-18', '1984-02-29',
        '2013-06-09', '2010-04-02', '2050-01-01', '2012-04-32', '2000-07-01',
        '1957-11-11', '1969-04-11', '1994-11-14', '2013-12-23', '1950-08-20',
        '1971-09-02', '1953-10-18', '1992-11-27', '2025-1040', '1941-11-07'
    ],
    'Género': [
        'X', 'F', 'M', 'M', 'M',
        'M', 'M', 'M', 'XM', 'M',
        'F', None, 'X', 'F', 'X',
        'F', 'M', 'F', 'X', 'M'
    ],
    'Estado Civil': [
        'divorciado', 'soltero', 'separado', 'diorciado', 'divorciado',
        'sotero', 'viudo', 'viudo', 'viudo', 'viudo',
        'viudo', 'casado', 'casado', 'separado', 'casado',
        'soltero', 'viud', 'separado', 'viudo', 'casado'
    ],
    'Dpto': [
        'Sucre', 'córdba', 'córdoba', 'antioquia', 'Cundinamarca',
        'Bolívar', 'antioquia', None, 'antioquia', 'córdoba',
        'antioquia', 'Atlántico', 'Sucre', 'córdoba', 'antioquia',
        'córdoba', 'antioquia', 'antioquia', 'córdoba', 'Bolívar'
    ],
    'Ciudad': [
        'Sincelej', 'lorica', 'ciudad_invalida', 'itagüí', 'bogotá',
        None, 'medellín', 'sahagún', 'medellín', 'sahagún',
        'itagüí', 'Barranquilla', 'San Marcos', 'sahagún', 'itagüí',
        'sahagún', 'medellín', 'itagüí', 'sahagún', 'Cartagena de Indias'
    ]
})

elif uploaded_file:
    df = pd.read_excel(uploaded_file)
else:
    df = None

if df is not None:
    st.markdown('---')
    st.subheader('Vista Previa')
    st.markdown("Se muestran las primeras 10 filas del archivo cargado para verificación rápida.")
    st.dataframe(df.head(10)) # Se muestran las primeras 10 filas del dataframe
    st.subheader("⚙️ Configura las validaciones")
    st.markdown("""
    Selecciona las columnas correspondientes a cada tipo de dato.  
    No es obligatorio llenar todas; puedes elegir solo las que quieras analizar.
    """)
    DEFAULT_OPTION = 'Select an option'
    options = [DEFAULT_OPTION] + list(df.columns)
    #Definimos  3 columnas para organizar mejor el espacio
    col1, col2, col3 = st.columns(3)
    with col1:
        nombre_col = st.selectbox('🏷️ Primer Nombre:', options)
        telefono_col = st.selectbox('📱 Teléfono:', options)
        tipo_doc_col = st.selectbox('🪪 Tipo de Documento:', options)
    with col2:
        apellido_col = st.selectbox('📝 Primer Apellido:', options)
        correo_col = st.selectbox('✉️ Email:', options) 
        doc_col = st.selectbox('🆔ID :', options)
    with col3:
        fecha_col = st.selectbox('🎂 Fecha de Nacimiento:', options)
        genero_col = st.selectbox('🚻 Género:', options)
        estado_civil_col = st.selectbox('💑 Estado Civil:', options)
    with st.expander("📍 Ubicación"):
        departamento_col = st.selectbox('🗺️ Departamento:', options)
        ciudad_col = st.selectbox('🏙️ Ciudad:', options)
    if st.button('🚀 Validar Datos'):
        resultados = [] #Creamos una lista para almacenar los resultados de la validación
        errores = {} #Diccionario para contar errores por columna. columna: numero de errores
        for idx, row in df.iterrows(): #Se obtiene el numero de la fila y los datos asociado a ella.
            fila = {'Fila': idx}
            # Validaciones solo si la columna fue seleccionada
            if nombre_col != DEFAULT_OPTION:
                nombre = row[nombre_col]
                nombre_valido = validate_Names(nombre) # True o False
                errores[nombre_col] = errores.get(nombre_col, 0) + (0 if nombre_valido else 1)
                fila['Nombre'] = nombre
                fila['is_name_valid'] = nombre_valido
            if apellido_col != DEFAULT_OPTION:
                apellido = row[apellido_col]
                apellido_valido = validate_Names(apellido)
                errores[apellido_col] = errores.get(apellido_col, 0) + (0 if apellido_valido else 1)
                fila['Apellido'] = apellido
                fila['is_lastname_valid'] = apellido_valido
            if telefono_col != DEFAULT_OPTION:
                tel = row[telefono_col]
                tel_valido = validate_phone(tel)
                errores[telefono_col] = errores.get(telefono_col, 0) + (0 if tel_valido else 1)
                fila['Teléfono'] = tel
                fila['is_phone_valid'] = tel_valido
            if correo_col != DEFAULT_OPTION:
                email = row[correo_col]
                email_valido = validate_Email(email)
                errores[correo_col] = errores.get(correo_col, 0) + (0 if email_valido else 1)
                fila['Correo'] = email
                fila['is_email_valid'] = email_valido
            if tipo_doc_col != DEFAULT_OPTION:
                tipo_doc = row[tipo_doc_col]
                fecha = row[fecha_col] if fecha_col != DEFAULT_OPTION  else None
                tipo_doc_valido = Validate_Document_type(tipo_doc, fecha)
                errores[tipo_doc_col] = errores.get(tipo_doc_col, 0) + (0 if tipo_doc_valido else 1)
                fila['Tipo de documento'] = tipo_doc
                fila['is_document_type_valid'] = tipo_doc_valido
            if doc_col != DEFAULT_OPTION:
                doc = row[doc_col]
                doc_valido = validate_ID(doc)
                errores[doc_col] = errores.get(doc_col, 0) + (0 if doc_valido else 1)
                fila['Documento'] = doc
                fila['is_document_valid'] = doc_valido
            if fecha_col != DEFAULT_OPTION:
                fecha = row[fecha_col]
                fecha_valida = validate_Birthday(fecha)
                errores[fecha_col] = errores.get(fecha_col, 0) + (0 if fecha_valida else 1)
                fila['Fecha de nacimiento'] = fecha
                fila['is_birthdate_valid'] = fecha_valida
            if genero_col != DEFAULT_OPTION:
                genero = row[genero_col]
                genero_valido = validate_Gender(genero)
                errores[genero_col] = errores.get(genero_col, 0) + (0 if genero_valido else 1)
                fila['Género'] = genero
                fila['is_gender_valid'] = genero_valido
            if estado_civil_col != DEFAULT_OPTION:
                estado_civil = row[estado_civil_col]
                estado_civil_valido = validate_marital_status(estado_civil)
                errores[estado_civil_col] = errores.get(estado_civil_col, 0) + (0 if estado_civil_valido else 1)
                fila['Estado civil'] = estado_civil
                fila['is_marital_status_valid'] = estado_civil_valido
            if departamento_col != DEFAULT_OPTION:
                departamento = row[departamento_col]
                departamento_valido = str(departamento).strip().lower() in departamento_ciudad_dict
                errores[departamento_col] = errores.get(departamento_col, 0) + (0 if departamento_valido else 1)
                fila['Departamento'] = departamento
                fila['is_state_valid'] = departamento_valido
            if ciudad_col != DEFAULT_OPTION :
                ciudad = row[ciudad_col]
                departamento = row[departamento_col] if departamento_col != DEFAULT_OPTION else None
                ciudad_valido = validate_Location(departamento, ciudad) # True o False
                errores[ciudad_col] = errores.get(ciudad_col, 0) + (0 if ciudad_valido else 1)
                fila['Ciudad'] = ciudad
                fila['is_city_valid'] = ciudad_valido
            resultados.append(fila)
            
        st.markdown('---')
        st.subheader('Resumen de Errores')
        st.dataframe(pd.DataFrame(resultados), hide_index=True) # Mostrar resultados de la validación
        st.subheader("📉 Recuento de errores por columna")
        st.markdown("Cantidad total de errores encontrados en cada campo validado.")
        st.dataframe(pd.DataFrame.from_dict(errores, orient='index', columns=['Errores']), hide_index=False)

        #Encontramos las columnas que no son nulas para calcular las metricas de calidad
        columnas_nulo = [col for col in [nombre_col, apellido_col, telefono_col, correo_col, tipo_doc_col, doc_col, fecha_col, genero_col, estado_civil_col, departamento_col, ciudad_col] if col != DEFAULT_OPTION ]
        total_nulos = df[columnas_nulo].isna().sum().sum()
        total_registros = df[columnas_nulo].shape[0] * len(columnas_nulo)
        total_errores = sum([errores.get(col,0) for col in errores])
        total_validados = total_registros - total_errores
        calidad = round((total_validados/total_registros)*100,2) if total_registros > 0 else 0
        
        st.subheader('Métricas de Calidad')
        colA, colB, colC, colD, colE = st.columns(5)
        with colA:
            st.metric(label="Datos", value=total_registros)
        with colB:
            st.metric(label="Errores", value=total_errores)
        with colC:
            st.metric(label="Validados", value=total_validados)
        with colD:
            st.metric(label="Calidad", value=f"{calidad}%")
        with colE:
            st.metric(label="Nulos", value=total_nulos)
        
        st.info(""" ✅ Validación completada correctamente.Puedes descargar el archivo Excel con los errores resaltados en **rojo** para corregirlos fácilmente.""")

        # Generar archivo Excel con errores resaltados 
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hoja de datos con errores
            df_export = df.copy()
            # Insertar la columna __row_idx__ al principio
            #df_export.insert(0, '__row_idx__', range(len(df_export)))
            # Resaltar errores en rojo
            error_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            df_export.to_excel(writer, index=False, sheet_name='Datos validados')
            workbook = writer.book
            worksheet = writer.sheets['Datos validados']
            # Mapear columnas y claves de validación
            valid_keys = {
                nombre_col: 'is_name_valid',
                apellido_col: 'is_lastname_valid',
                telefono_col: 'is_phone_valid',
                correo_col: 'is_email_valid',
                tipo_doc_col: 'is_document_type_valid',
                doc_col: 'is_document_valid',
                fecha_col: 'is_birthdate_valid',
                genero_col: 'is_gender_valid',
                estado_civil_col: 'is_marital_status_valid',
                departamento_col: 'is_state_valid',
                ciudad_col: 'is_city_valid'
            }
            for i, res in enumerate(resultados):
                for col, valid_key in valid_keys.items():
                    if valid_key in res and not res[valid_key]:
                        cell = worksheet.cell(row=i+2, column=list(df_export.columns).index(col)+1)
                        cell.fill = error_fill
        output.seek(0)
        
        st.download_button('Resultados Validados', data=output, file_name='Resultados.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

   

